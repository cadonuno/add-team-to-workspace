import sys
import requests
import getopt
import json
import urllib.parse
from veracode_api_signing.plugin_requests import RequestsAuthPluginVeracodeHMAC
import openpyxl
import time

from veracode_api_signing.credentials import get_credentials


headers = {
    "User-Agent": "Adding teams to workspaces - python script",
    "Content-Type": "application/json"
}

failed_attempts = 0
max_attempts_per_request = 10
sleep_time = 10

def print_help():
    """Prints command line options and exits"""
    print("""add-team-to-workspace.py -f <excel_file_with_workspaces_and_teams> [-d] [--use_team_id]"
        Reads all lines in <excel_file_with_workspaces_and_teams>, for each line, it will search for the column
        *Can optionally use team id's instead of names
""")
    sys.exit()

def get_workspace_by_name(api_base, workspace_name, verbose):
    global failed_attempts
    global sleep_time
    global max_attempts_per_request
    path = f"{api_base}srcclr/v3/workspaces?filter%5Bworkspace%5D={urllib.parse.quote(workspace_name, safe='')}"
    if verbose:
        print(f"Calling: {path}")

    response = requests.get(path, auth=RequestsAuthPluginVeracodeHMAC(), headers=headers)
    data = response.json()
    
    if response.status_code == 200:
        if verbose:
            print(data)
        if len(data["_embedded"]["workspaces"]) > 0:
            return find_exact_match(data["_embedded"]["workspaces"], workspace_name, "name")
        else:
            print(f"ERROR: No workspace named '{workspace_name}' found")
            return f"ERROR: No workspace named '{workspace_name}' found"
    else:
        print(f"ERROR: trying to get workspace named {workspace_name}")
        print(f"ERROR: code: {response.status_code}")
        print(f"ERROR: value: {data}")
        failed_attempts+=1
        if (failed_attempts < max_attempts_per_request):
            time.sleep(sleep_time)
            return get_workspace_by_name(api_base, workspace_name, verbose)
        else:
            return f"ERROR: trying to get workspace named {workspace_name}"
    

def get_team_by_name(api_base, team_name, verbose):
    global failed_attempts
    global sleep_time
    global max_attempts_per_request
    path = f"{api_base}api/authn/v2/teams?all_for_org=true&team_name={urllib.parse.quote(team_name, safe='')}"
    if verbose:
        print(f"Calling: {path}")

    response = requests.get(path, auth=RequestsAuthPluginVeracodeHMAC(), headers=headers)
    data = response.json()

    if response.status_code == 200:
        if verbose:
            print(data)
        if len(data["_embedded"]["teams"]) > 0:
            return find_exact_match(data["_embedded"]["teams"], team_name, "team_name")
        else:
            print(f"ERROR: No team named '{team_name}' found")
            return f"ERROR: No team named '{team_name}' found"
    else:
        print(f"ERROR: trying to get team named {team_name}")
        print(f"ERROR: code: {response.status_code}")
        print(f"ERROR: value: {data}")
        failed_attempts+=1
        if (failed_attempts < max_attempts_per_request):
            time.sleep(sleep_time)
            return get_team_by_name(api_base, team_name, verbose)
        else:
            return f"ERROR: trying to get team named {team_name}"

def add_team_to_workspace(api_base, workspace_name, team_parameter_value, verbose, use_team_id):
    global failed_attempts
    global sleep_time
    global max_attempts_per_request
    print(f"Looking for workspace: {workspace_name}")
    workspace = get_workspace_by_name(api_base, workspace_name, verbose)
    if workspace:
        workspace_guid = workspace["id"]
    else:
        print(f"ERROR: No workspace named '{workspace_name}' found")
        return f"ERROR: No workspace named '{workspace_name}' found"

    if use_team_id:
        if verbose:
            print(f"Using team id of {team_parameter_value}")
        team_id=team_parameter_value
    else:
        print(f"Looking for team: {team_parameter_value}")
    
        team = get_team_by_name(api_base, team_parameter_value, verbose);
        if verbose:
            print(team)
        if team:
            team_id = team["team_legacy_id"]
        else:
            print(f"ERROR: No team named '{team_parameter_value}' found")
            return f"ERROR: No team named '{team_parameter_value}' found"
    
    path = f"{api_base}srcclr/v3/workspaces/{workspace_guid}/teams/{team_id}"

    if verbose:
        print(f"Calling {path}")

    response = requests.put(path, auth=RequestsAuthPluginVeracodeHMAC(), headers=headers)

    if verbose:
        print(f"status code {response.status_code}")
        
    if response.status_code == 204:
        print("Successfully added team to workspace.")
        return "done"
    else:
        print(f"Unable to provide access to team: {response.status_code}")
        if verbose and response:
            body = response.json()
            if body:
                print(body)
        failed_attempts+=1
        if (failed_attempts < max_attempts_per_request):
            time.sleep(sleep_time)
            return add_team_to_workspace(api_base, workspace_name, team_parameter_value, verbose, use_team_id)
        else:
            return f"Unable to provide access to team: {response.status_code}"

def find_exact_match(list, to_find, field_name):
    for index in range(len(list)):
        if list[index][field_name] == to_find:
            return list[index]
    print(f"Unable to find a member of list with {field_name} equal to {to_find}")
    raise NoExactMatchFoundException(f"Unable to find a member of list with {field_name} equal to {to_find}")
        
def get_api_base():
    api_key_id, api_key_secret = get_credentials()
    api_base = "https://api.veracode.{instance}/"
    if api_key_id.startswith("vera01"):
        return api_base.replace("{instance}", "eu", 1)
    else:
        return api_base.replace("{instance}", "com", 1)

def main(argv):
    """Simple command line support for bulk adding teams to SCA workspaces"""
    global failed_attempts
    try:
        verbose = False
        file_name = ''
        use_team_id = False

        opts, args = getopt.getopt(argv, "hdf:", ["file_name=", "use_team_id"])
        for opt, arg in opts:
            if opt == '-h':
                print_help()
            if opt == '-d':
                verbose = True
            if opt in ('-f', '--file_name'):
                file_name=arg
            if opt == '--use_team_id':
                use_team_id = True


        api_base = get_api_base()
        if file_name:
            print(file_name)

            excel_file = openpyxl.load_workbook(file_name)
            excel_sheet = excel_file.active
            
            for row in range(2, excel_sheet.max_row):          
                print(f"Starting proccess for row {row}:")
                failed_attempts = 0
                workspace=excel_sheet.cell(row = row, column = 1).value
                team=excel_sheet.cell(row = row, column = 2).value
                status=excel_sheet.cell(row = row, column = 3).value
                print(f"Found: {workspace} | {team} | {status}")
                if (status == 'done'):
                    print("Skipping row as it was already done")
                else:
                    try:
                        status=add_team_to_workspace(api_base, workspace, team, verbose, use_team_id)
                    except NoExactMatchFoundException:
                        status= NoExactMatchFoundException.get_message()
                    excel_sheet.cell(row = row, column = 3).value=status

        else:
            print_help()
    except requests.RequestException as e:
        print("An error occurred!")
        print(e)
        sys.exit(1)
    finally:
        excel_file.save(filename=file_name)


if __name__ == "__main__":
    main(sys.argv[1:])

class NoExactMatchFoundException(Exception):
    message=""
    def __init__(self, message_to_set):
        message = message_to_set

    def get_message(self):
        return message